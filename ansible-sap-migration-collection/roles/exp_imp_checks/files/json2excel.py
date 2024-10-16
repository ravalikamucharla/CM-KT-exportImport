#!/usr/bin/python3
import json
import os
import sys
import warnings
import glob
import copy
from datetime import datetime
import re
import argparse
from os import path
import subprocess
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
from openpyxl.styles import Font, NamedStyle,Border,Side,PatternFill
from flatten_json import flatten
import unicodedata

filecnt = 0

class jsonToExcel():
    def __init__(self):
        self.sheet_name=""
        self.workbook_name = ""
        self.flag=1
        self.source_path=""
        self.mini_check_args=""
        self.bold_style=""
        self.redFill = PatternFill(start_color='00FFF000',
                              end_color='00FFF000',
                              fill_type='solid')
        self.normal_cell_style=""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheet=self.workbook.create_sheet(title=self.sheet_name)
        self.lst_jsons=[]

    def findnth(self,inputstr, substr, n):
        start = inputstr.find(substr)
        while start >= 0 and n > 1:
            start = inputstr.find(substr, start + len(substr))
            n -= 1
        return start

    def countX(self,lst, x):
        return lst.count(x)

    def replace_control_characters(self,s):
        if not type(s) == str:
            return str(s)
        matched = s
        lstf=[]
        for c in s:
            if unicodedata.category(c)[0] == 'C':
                search = re.findall(r'\x1b', s)
                if search:
                    matched=bytes(matched,'utf-8')
                    break
        return matched
        return str(matched)

    def get_proper_flatten(self,flattened,lstkey=None):
        '''
        This function prefixes key in json
        '''
        new_json={}
        cpflatten=copy.deepcopy(flattened)
        if lstkey:
            for eachk,eachv in cpflatten.items():
                new_json[lstkey+"_"+eachk]=eachv
        else:
            return cpflatten
        return new_json

    def parse_args(self,argvars):
        try:

            arg_len = len(argvars)
            if arg_len > 0:
                json_path = argvars[0]
                if not os.path.isabs(json_path):
                    json_path = os.path.join(os.getcwd(), json_path)
                self.source_path=json_path
                outpath=argvars[1]
                self.workbook_name = os.path.join(outpath, "checks.xlsx")

            else:
                print("Please supply 'Input path' as argument")
        except Exception as err:
            print("ERROR :",err)



    def def_styles(self):
        bold = NamedStyle(name="Bold")
        bold.font = Font(color='00000000', bold=True)
        bold.fill = self.redFill
        bold.alignment.wrap_text = True
        self.bold_style = bold
        normal_txt = NamedStyle(name="normaltext")
        normal_txt.border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))
        normal_txt.alignment.wrap_text = True
        self.normal_cell_style=normal_txt

    def read_input(self):
        rootdir = self.source_path
        lstdirs=[]
        files_flag=0
        curdirfile = [x for x in os.listdir(rootdir) if x.endswith(".json")]
        all_dirs = [x[0] for x in os.walk(rootdir)]
        for file in all_dirs:
            if not file == '.':
                d = os.path.join(rootdir, file)
                if os.path.isdir(d):
                    lstdirs.append(d)

        if curdirfile:
            files_flag =self.source_path
        return lstdirs,files_flag


    def set_col_width(self,lstheader):
        cols=len(lstheader)
        for each_col in range(cols):
            self.sheet.column_dimensions[get_column_letter(each_col+1)].width=30


    def get_headers(self,header_data):
        lsthead=[]
        for each in header_data:
            if "_" in each:
                search = re.findall(r'_[\d]+_', each)
                if search == []:
                    if each not in lsthead:
                        lsthead.append(each)
                else:
                    concat_str=""
                    cpeach=copy.deepcopy(each)
                    for eachs in search:
                        splts = cpeach.split(eachs,1)
                        cpeach=splts[1]
                        concat_str=concat_str+"_"+splts[0]
                    concat_str = concat_str + "_" + splts[1]
                    item=concat_str[1:]
                    if item not in lsthead:
                        lsthead.append(item)
            else:
                if each not in lsthead:
                    lsthead.append(each)
        return lsthead

    def merge_headers(self,cur_header,new_header):
        merged = copy.deepcopy(cur_header)
        newly_added=[]
        for each in new_header:
            if not each in merged:
                merged.append(each)
                newly_added.append(each)
        return merged,newly_added

    def create_json_table(self,flatten_data,header):
        findata={}
        sepcol=[]
        maxno=0
        lstnos=[]
        maxnew=0
        for eachk in flatten_data.keys():
            maxnew = 0
            if "_" in eachk:
                search=re.findall(r'_[\d]+_', eachk)
                for eachno in search:
                    searchn = re.findall(r'[\d]+', eachno)
                    if searchn[0]:
                        conv = searchn[0]
                        lstnos.append(int(conv))
                if lstnos:
                    maxnew=max(lstnos)
            if maxno < maxnew :
                maxno = maxnew
        for x in range(maxno+1):
            findata[x]=[]
        for ind,each in enumerate(header):
            lstdat=[]
            for eachk in flatten_data.keys():
                cpeachk=copy.deepcopy(eachk)
                if "_" in eachk:
                    matchstr = re.findall(r'_[\d]+_', eachk)
                    if matchstr:
                        matched = re.sub(r'_[\d]+_', "_",cpeachk)
                    else:
                        matched =eachk
                    if each == matched:
                        lstdat.append(eachk)
                else:
                    if eachk == each:
                        if eachk not in lstdat:
                            lstdat.append(eachk)
            matchcheck = re.findall(r'_[\d]+_', lstdat[0])
            subs = copy.deepcopy(lstdat[0])
            if not matchcheck:
                lstitem = lstdat[0]
                for index in range(maxno+1):
                    findata[index].append({lstitem: flatten_data[lstdat[0]]})
            else:
                index=0
                lstitem=re.sub(r'_[\d]+_', "_",subs)
                for each_lst in lstdat:
                    findata[index].append({lstitem:flatten_data[each_lst]})
                    index +=1
                while index <= maxno:
                    findata[index].append({lstitem: ""})
                    index +=1
        return findata

    def merge_flatten_tables(self,finflatten,p1,header):
        flat_tab=copy.deepcopy(finflatten)
        keylst=flat_tab.keys()
        maxind=max(keylst)
        max_p1=0
        maxnew=0
        for ind,each in enumerate(header):
            maxnew = 0
            for eachk in p1.keys():
                lstnos=[]
                if "_" in eachk:
                    search = re.findall(r'_[\d]+_', eachk)
                    for eachno in search:
                        searchn = re.findall(r'[\d]+', eachno)
                        if searchn[0]:
                            conv = searchn[0]
                            lstnos.append(int(conv))
                if lstnos:
                    maxnew = max(lstnos)
                if max_p1 < maxnew:
                    max_p1 = maxnew
        for count in range(maxind+1,max_p1+maxind+2):
            flat_tab[count]=[]
        count = maxind+1
        for index,each_head in enumerate(header):
            val_flag=0
            lstdata=[]
            for one_entry in p1.keys():
                if "_" in one_entry:
                    cpentry=copy.deepcopy(one_entry)
                    find_digit=re.findall(r'_[\d]_+', one_entry)
                    if find_digit:
                        spltsearch = re.sub(r'_[\d]+_', "_",cpentry)
                    else:
                        spltsearch =cpentry
                    if each_head == spltsearch:
                        lstdata.append(one_entry)
                else:
                    if one_entry == each_head:
                        if one_entry not in lstdata:
                            lstdata.append(one_entry)
            if len(lstdata) > (max_p1+1):
                diff=len(lstdata)-(max_p1+1)
                cval= len(flat_tab)
                for lpvar in range(diff):
                    flat_tab[cval]=[]
                    cval +=1
            if lstdata == []:
                for count in range(maxind + 1, max_p1 + maxind + 2):
                    flat_tab[count].append({each_head:""})

            elif len(lstdata) < (max_p1+1):
                itemk=copy.deepcopy(lstdata[0])
                find_st = re.findall(r'_[\d]_+', itemk)
                if find_st:
                    spltsea = re.sub(r'_[\d]+_', "_", itemk)
                    cntlp = maxind + 1
                    cntlpf = len(lstdata)
                    lpcou = 0
                    while (lpcou < len(lstdata)):
                        flat_tab[cntlp].append({spltsea: p1[lstdata[lpcou]]})
                        cntlp += 1
                        lpcou += 1
                    while (cntlp < (max_p1 + maxind + 2)):
                        flat_tab[cntlp].append({spltsea: ""})
                        cntlp += 1
                else:
                    spltsea=itemk
                    for cp in range(maxind + 1,max_p1 + maxind + 2):
                        flat_tab[cp].append({spltsea:p1[lstdata[0]] })
            else:
                lendat = len(lstdata)
                count=maxind+1
                counter =len(flat_tab)-1
                count=maxind + 1
                for each_lst in lstdata:
                    if flat_tab[count]:
                        flat_tab[count].append({each_lst: p1[each_lst]})
                    else:
                        flat_tab[count]=[]
                        flat_tab[count].append({each_lst: p1[each_lst]})
                    count +=1
        return flat_tab

    def write_invalid_json(self,filename,row,col):
        col += 2
        self.sheet.cell(row, col, "Invalid JSON")
        self.format_table_title(row)

    def format_table_title(self,row):
        sel_row = self.sheet[row:row]
        for cell in sel_row:
            cell.style = self.bold_style
        if row == 2:
            sel_row = self.sheet[row:row]
            for cell in sel_row:
                cell.style = self.bold_style

    def add_blank_columns_to_existing_table(self,json_table,new_headers):
        cur_json=copy.deepcopy(json_table)
        keylst=list(cur_json.keys())
        max_elmt=max(keylst)
        for each_head in new_headers:
            for count in range(max_elmt+1):
                cur_json[count].append({each_head:""})
        return cur_json
    def flatten_many_jsons_and_merge(self,lstdic,lstkey=None):
        '''
        Get list of jsons and flatten and merge
        '''
        cur_header =[]
        for indd, each1 in enumerate(lstdic):
            p1 = flatten(each1)
            p1=self.get_proper_flatten(p1,lstkey)
            tmpheader = self.get_headers(p1.keys())
            cur_header, new_header_items = self.merge_headers(cur_header, tmpheader)
            if indd ==0:
                finflatten = self.create_json_table(p1, cur_header)
            if indd > 0:
                finflatten = self.add_blank_columns_to_existing_table(finflatten,new_header_items)
            if indd > 0 :
                finflatten = self.merge_flatten_tables(finflatten, p1, cur_header)
        return  finflatten,cur_header

    def format_flat_json(self,flat_json):
        ft_json=copy.deepcopy(flat_json)
        new_ft_json={}
        for eachkey,eachval in ft_json.items():
            nkey=eachkey
            if "_" in eachkey:
                find_digit = re.findall(r'_[\d]_+', eachkey)
                if find_digit:
                    comb=""
                    recsplt=copy.deepcopy(eachkey)
                    for each_search in find_digit:
                        spltsearch = recsplt.split(each_search, 1)
                        recsplt=spltsearch[1]
                        comb=comb+"_"+spltsearch[0]
                    comb=comb+"_"+spltsearch[1]
                    nkey=comb[1:]
            new_ft_json[nkey]=eachval
        return new_ft_json

    def single_json_flatten_and_create_table(self,data):
        flat_json = flatten(data)
        lstheader = self.get_headers(flat_json.keys())
        json_table = self.create_json_table(flat_json, lstheader)
        return json_table,lstheader

    def write_table_header(self,lstheader,row):
        row = row + 1
        for ind, each_head in enumerate(lstheader):
            self.sheet.cell(row, ind + 1, each_head)
        sel_row = self.sheet[row:row]
        for cell in sel_row:
            cell.style = self.normal_cell_style
        return row

    def write_table_data(self,json_table,lstheader,row):
        col = 0
        row += 1
        counter = 0
        tot_jsonlen=len(json_table)
        check=1
        countm=0
        for eachdat, eachval in json_table.items():
            if tot_jsonlen > 5000:
                countm +=1
                if countm == 500 * check:
                    check = check + 1
            for ind, eachvalues in enumerate(eachval):
                writekey = lstheader[ind]
                valuek = list(eachvalues.keys())[0]
                value_to_write = eachvalues[valuek]
                self.sheet.cell(row, ind + 1, str(self.replace_control_characters(value_to_write)))
                sel_row = self.sheet[row:row]
                for cell in sel_row:
                    cell.style = self.normal_cell_style
            row += 1
        return row
    def write_to_excel(self,each_dir):
        global filecnt
        try:

            self.lst_jsons = glob.glob(os.path.join(self.source_path,each_dir, "*.json"))
            self.sheet_name=os.path.basename(each_dir)
            row = 1
            col = 1
            self.sheet = self.workbook.create_sheet(title=self.sheet_name)
            row = 1
            # Iterate each json file and read
            rowcheck=0
            for index, each in enumerate(self.lst_jsons):
                filename = os.path.basename(each)
                print("%s. %s " %(index+1 ,filename))
                # For seperate excel sheets
                row +=1
                col = 1

                # Read json
                try:
                    with open(each,encoding='utf-8') as json_file:
                        data = json.load(json_file,strict=False)
                except Exception as err:
                    print(json.dumps({"Error": "Error reading json file %s, Error:%s" % (filename, err)}))
                    continue

                #Check empty json
                cpdata=copy.deepcopy(data)
                if type(cpdata) == list :
                    dic_val = None
                    self.sheet.cell(row, col, '%s' % (filename.split(".json")[0]))
                    json_table, lstheader = self.flatten_many_jsons_and_merge(cpdata, dic_val)
                else:
                    empty_keys = [key for key, val in data.items() if not val]
                    empty_dict = False
                    if len(empty_keys) == len(data.keys()):
                        empty_dict = True

                    #Write json file name to current row and apply bold style
                    self.sheet.cell(row, col, '%s' % (filename.split(".json")[0]))
                    # If empty dict, write in excel as "invalid json
                    if empty_dict:
                        self.write_invalid_json(filename,row,col)
                        #Move to next json
                        continue

                    #Check json format and flatten seperately if needed
                    dic_val = list(data.keys())[0]
                    lstdic = data[dic_val]

                    if len(lstdic) < 1:
                        self.write_invalid_json(filename,row,col)
                        continue
                    if type(lstdic) == dict or type(lstdic) == list:
                        #If master key has only one list item, its needed to flatten inner data seperately
                        if type(lstdic) == list:
                            cur_header = []  # To merge headers, initializing current header as blank
                            json_table,lstheader  = self.flatten_many_jsons_and_merge(lstdic,dic_val)
                    else:
                        json_table,lstheader = self.single_json_flatten_and_create_table(data)

                #Start writing to table
                self.format_table_title(row)
                self.set_col_width(lstheader)
                # write table header
                row = self.write_table_header(lstheader,row)
                #write table data
                row = self.write_table_data(json_table,lstheader,row)
            return
        except Exception as exceptn:
            today = datetime.now()
            today = today.strftime('%Y%m%d_%I%M%S%p')
            filewrite = self.workbook_name.split(".xlsx")[0] + '_%s.xlsx' % (today)
            self.workbook.save(filewrite)
            print(json.dumps({"ERROR": exceptn}, default=str))
            pass
    def write_json(self,data,wpath):
        try:
            if not path.exists(os.path.join(wpath,"Hana_Mini_Check")):
                os.makedirs(os.path.join(wpath,"Hana_Mini_Check"))
            with open(os.path.join(wpath,"Hana_Mini_Check","hana_db_check.json"),'w') as ofile:
                json.dump(data,ofile)
            return
        except Exception as err:
            print(json.dumps({"ERROR": err}, default=str))
            pass
    def write_hana_db_checks(self):
        try:
            cmd = ["python", "health_check.py", self.mini_check_args]
            print("CMD to call",cmd)
            output = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
            out = json.loads(str(output.stdout.read(), 'utf-8'))
            self.write_json(out,self.source_path)
            return
        except Exception as err:
            print(json.dumps({"ERROR": err}, default=str))
            pass

try:
    json_to_excel = jsonToExcel()

    parser = argparse.ArgumentParser("json_to_excel")
    cur_path = os.path.join(os.getcwd(), "Output")

    parser.add_argument('--input', '-i', help='Input path')
    parser.add_argument('--output', '-o', default= cur_path,help='Optional output path')
    args = parser.parse_args()
    if not path.exists(args.output):
        os.makedirs(args.output)
    argvars=[args.input,args.output]
    json_to_excel.parse_args(argvars)
    json_to_excel.def_styles()
    today = datetime.now()
    today = today.strftime('%Y%m%d_%I%M%S')
    lst_ips,has_files = json_to_excel.read_input()
    json_to_excel.workbook.remove(json_to_excel.workbook.active)
    for eachdir in lst_ips[1:]:
        print("Processing folder -%s" %eachdir)
        json_to_excel.write_to_excel(eachdir)
    if has_files:
        json_to_excel.write_to_excel(has_files)
    filewrite = json_to_excel.workbook_name.split(".xlsx")[0] + '_%s.xlsx' % (today)
    json_to_excel.workbook.save(filewrite)
except IOError as err:
    today = datetime.now()
    today = today.strftime('%Y%m%d_%I%M%S%p')
    print(json.dumps({"Error":"Please close the output file if its open, Error:%s" %err}))
    filewrite = json_to_excel.workbook_name.split(".xlsx")[0]+'_%s.xlsx'  %(today)
    json_to_excel.workbook.save(filewrite)
except Exception as err:
    today = datetime.now()
    today = today.strftime('%Y%m%d_%I%M%S%p')
    print(json.dumps({"Error": err}, default=str))
    filewrite = json_to_excel.workbook_name.split(".xlsx")[0]+'_%s.xlsx'  %(today)
    json_to_excel.workbook.save(filewrite)
